import os
from fastapi import APIRouter, Depends, Query, HTTPException, Security, Request
from fastapi.security import APIKeyQuery
from fastapi.security.api_key import APIKeyHeader
from sqlalchemy.orm import Session
from sqlalchemy import func, case
from typing import Optional, List
from datetime import datetime, timezone, timedelta
from pydantic import BaseModel
import io

from fastapi.responses import StreamingResponse
from database import get_db
from models.scan_log import ScanLog
from models.admin_scan_log import AdminScanLog
from schemas.responses import ScanLogOut, AdminStats, UserScanSummary, UserDetailsOut, AdminScanRecord, AdminMonitorScan
from main import limiter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

router = APIRouter()

# ── Admin API-Key guard ────────────────────────────────────────────────────────
# If ADMIN_API_KEY is set in .env this is enforced on every /admin route.
# If it is NOT set the check is skipped (backward-compatible — shows a warning
# in the startup log instead).

_ADMIN_API_KEY   = os.getenv("ADMIN_API_KEY", "")
_api_key_header  = APIKeyHeader(name="X-Admin-Key", auto_error=False)
_api_key_query   = APIKeyQuery(name="api_key", auto_error=False)

VALID_FEATURES = {"url_scan", "screenshot_scan", "qr_scan", "otp_scan", "upi_scan", "voice_scan"}
VALID_VERDICTS = {"SAFE", "SUSPICIOUS", "DANGEROUS"}

async def verify_admin(
    api_key_header: Optional[str] = Security(_api_key_header),
    api_key_query: Optional[str] = Security(_api_key_query),
):
    """Dependency — enforces X-Admin-Key when ADMIN_API_KEY env var is configured."""
    if not _ADMIN_API_KEY:
        # Key not configured: allow through (backward-compat dev mode)
        return
    api_key = api_key_header or api_key_query
    if api_key != _ADMIN_API_KEY:
        raise HTTPException(status_code=401, detail="Invalid or missing X-Admin-Key header or api_key parameter.")


# ── Routes ────────────────────────────────────────────────────────────────────

@router.get("/scans", response_model=List[AdminScanRecord], dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_scans(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    feature: Optional[str] = None,
    verdict: Optional[str] = None,
    query: Optional[str] = None,
    db: Session = Depends(get_db),
):
    q = db.query(ScanLog)
    if feature and feature != "ALL":
        q = q.filter(ScanLog.feature == feature)
    if verdict and verdict != "ALL":
        q = q.filter(ScanLog.verdict == verdict)
    if query:
        search_pattern = f"%{query.strip().lower()}%"
        q = q.filter(
            func.lower(ScanLog.input_data).like(search_pattern) |
            func.lower(ScanLog.user_email).like(search_pattern) |
            func.lower(ScanLog.user_name).like(search_pattern)
        )
    q = q.order_by(ScanLog.scanned_at.desc())
    q = q.offset((page - 1) * per_page).limit(per_page)
    logs = q.all()

    records = []
    for l in logs:
        feat_clean = (l.feature or "unknown").replace("_scan", "").upper()
        records.append(AdminScanRecord(
            id=l.id,
            user_id=getattr(l, 'user_id', None) or f"usr-{l.id:04d}",
            user_name=l.user_name or "User",
            user_email=l.user_email or "user@cybershield.local",
            scan_type=feat_clean,
            result=l.verdict.title() if l.verdict else "Safe",
            confidence=l.confidence or 0.0,
            analysis=l.explanation or "Analysis completed cleanly.",
            created_at=l.scanned_at,
            status="completed"
        ))
    return records



@router.get("/logs", response_model=List[ScanLogOut], dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_logs(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=100),
    feature: Optional[str] = None,
    verdict: Optional[str] = None,
    query: Optional[str] = None,
    db: Session = Depends(get_db),
):
    # Allowlist filter values — prevent unexpected query parameters
    if feature and feature not in VALID_FEATURES:
        raise HTTPException(status_code=400, detail=f"Invalid feature. Allowed: {sorted(VALID_FEATURES)}")
    if verdict and verdict not in VALID_VERDICTS:
        raise HTTPException(status_code=400, detail=f"Invalid verdict. Allowed: {sorted(VALID_VERDICTS)}")

    q = db.query(ScanLog)
    if feature:
        q = q.filter(ScanLog.feature == feature)
    if verdict:
        q = q.filter(ScanLog.verdict == verdict)
    if query:
        search_pattern = f"%{query.strip().lower()}%"
        q = q.filter(
            func.lower(ScanLog.input_data).like(search_pattern) |
            func.lower(ScanLog.user_email).like(search_pattern) |
            func.lower(ScanLog.user_name).like(search_pattern)
        )
    q = q.order_by(ScanLog.scanned_at.desc())
    q = q.offset((page - 1) * per_page).limit(per_page)
    return q.all()


@router.get("/stats", response_model=AdminStats, dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_stats(request: Request, db: Session = Depends(get_db)):
    total = db.query(func.count(ScanLog.id)).scalar() or 0
    threats = db.query(func.count(ScanLog.id)).filter(
        ScanLog.verdict.in_(["DANGEROUS", "SUSPICIOUS"])
    ).scalar() or 0
    safe_rate = round(((total - threats) / total * 100), 1) if total > 0 else 100.0

    total_profiles = db.query(func.count(func.distinct(ScanLog.user_email))).scalar() or 0

    by_category_rows = (
        db.query(ScanLog.feature, func.count(ScanLog.id))
        .group_by(ScanLog.feature)
        .all()
    )
    by_category = {row[0]: row[1] for row in by_category_rows}

    today = datetime.now(timezone.utc).date()
    today_count = db.query(func.count(ScanLog.id)).filter(
        func.date(ScanLog.scanned_at) == today
    ).scalar() or 0

    daily_counts = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = db.query(func.count(ScanLog.id)).filter(
            func.date(ScanLog.scanned_at) == day
        ).scalar() or 0
        daily_counts.append({"date": day.isoformat(), "count": count, "label": day.strftime("%a")})

    return AdminStats(
        total=total,
        threats=threats,
        safe_rate=safe_rate,
        by_category=by_category,
        daily_counts=daily_counts,
        today_count=today_count,
        total_profiles=total_profiles,
    )


@router.get("/user/details", response_model=UserDetailsOut, dependencies=[Depends(verify_admin)])
@limiter.limit("60/minute")
async def get_user_details(request: Request, email: str = Query(...), db: Session = Depends(get_db)):
    email_clean = email.strip().lower()
    logs = (
        db.query(ScanLog)
        .filter(func.lower(ScanLog.user_email) == email_clean)
        .order_by(ScanLog.scanned_at.desc())
        .all()
    )

    if not logs:
        return UserDetailsOut(
            user_name="User",
            user_email=email_clean,
            total_scans=0,
            threats=0,
            safe_rate=100.0,
            last_scanned_at=None,
            created_at=None,
            by_category={},
            scan_history=[],
        )

    user_name = logs[0].user_name or "User"
    total_scans = len(logs)
    threats = sum(1 for l in logs if l.verdict in ["DANGEROUS", "SUSPICIOUS"])
    safe_rate = round(((total_scans - threats) / total_scans * 100), 1) if total_scans > 0 else 100.0
    last_scanned_at = logs[0].scanned_at
    created_at = logs[-1].scanned_at

    by_category = {}
    for l in logs:
        feat = l.feature or "unknown"
        by_category[feat] = by_category.get(feat, 0) + 1

    return UserDetailsOut(
        user_name=user_name,
        user_email=email_clean,
        total_scans=total_scans,
        threats=threats,
        safe_rate=safe_rate,
        last_scanned_at=last_scanned_at,
        created_at=created_at,
        by_category=by_category,
        scan_history=logs,
    )



class UpdateUserNameRequest(BaseModel):
    email: str
    new_name: str

@router.post("/user/update-name")
async def update_user_name(body: UpdateUserNameRequest, db: Session = Depends(get_db)):
    email_clean = body.email.strip().lower()
    name_clean = body.new_name.strip()
    if email_clean and name_clean:
        db.query(ScanLog).filter(func.lower(ScanLog.user_email) == email_clean).update(
            {ScanLog.user_name: name_clean}, synchronize_session=False
        )
        db.query(AdminScanLog).filter(func.lower(AdminScanLog.user_email) == email_clean).update(
            {AdminScanLog.user_name: name_clean}, synchronize_session=False
        )
        db.commit()
    return {"status": "ok", "email": email_clean, "new_name": name_clean}


@router.get("/users", response_model=List[UserScanSummary], dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_user_stats(request: Request, db: Session = Depends(get_db)):

    rows = (
        db.query(
            ScanLog.user_email,
            func.count(ScanLog.id).label("total_scans"),
            func.sum(
                case(
                    (ScanLog.verdict.in_(["DANGEROUS", "SUSPICIOUS"]), 1),
                    else_=0
                )
            ).label("threats"),
            func.max(ScanLog.scanned_at).label("last_scanned_at")
        )
        .group_by(ScanLog.user_email)
        .all()
    )

    results = []
    for r in rows:
        email = r.user_email or "user@cybershield.local"
        latest_log = (
            db.query(ScanLog.user_name)
            .filter(ScanLog.user_email == email)
            .order_by(ScanLog.scanned_at.desc())
            .first()
        )
        name = (latest_log[0] if latest_log and latest_log[0] else "User")


        total = r.total_scans or 0
        threats = r.threats or 0
        safe_count = total - threats
        safe_rate = round((safe_count / total * 100), 1) if total > 0 else 100.0

        results.append(UserScanSummary(
            user_name=name,
            user_email=email,
            total_scans=total,
            threats=threats,
            safe_rate=safe_rate,
            last_scanned_at=r.last_scanned_at
        ))

    return results



@router.get("/export/csv", dependencies=[Depends(verify_admin)])
@limiter.limit("20/minute")
async def export_csv(request: Request, db: Session = Depends(get_db)):
    logs = db.query(ScanLog).order_by(ScanLog.scanned_at.desc()).all()
    
    # Create an in-memory workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "CyberShield Logs"
    ws.views.sheetView[0].showGridLines = True

    # Styling definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="2F6EFF", end_color="2F6EFF", fill_type="solid") # Royal Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    safe_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    safe_font = Font(name=font_family, size=10, bold=True, color="375623")
    
    suspicious_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    suspicious_font = Font(name=font_family, size=10, bold=True, color="7F6000")
    
    dangerous_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    dangerous_font = Font(name=font_family, size=10, bold=True, color="C65911")
    
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    data_font = Font(name=font_family, size=10)
    
    # Headers
    headers = ["ID", "User Name", "User Email", "Feature", "Input Data", "Verdict", "Confidence", "Explanation", "Scanned At"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin

    # Fill data
    for row_idx, log in enumerate(logs, start=2):
        # Format scanned_at to date/time format: YYYY-MM-DD HH:MM:SS
        scanned_at_str = ""
        if log.scanned_at:
            scanned_at_str = log.scanned_at.strftime("%Y-%m-%d %H:%M:%S")

        row_data = [
            f"#LOG-{log.id:04d}" if isinstance(log.id, int) else str(log.id),
            str(log.user_name or "User"),
            str(log.user_email or "user@cybershield.local"),
            str(log.feature or "").replace("_scan", "").upper(),

            str(log.input_data or ""),
            str(log.verdict or ""),
            f"{log.confidence}%" if log.confidence is not None else "N/A",
            str(log.explanation or ""),
            scanned_at_str
        ]

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border_thin
            
            # Alignments
            if col_idx in [1, 4, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Verdict Column Color Styling
            if col_idx == 6:
                if val == "SAFE":
                    cell.fill = safe_fill
                    cell.font = safe_font
                elif val == "SUSPICIOUS":
                    cell.fill = suspicious_fill
                    cell.font = suspicious_font
                elif val == "DANGEROUS":
                    cell.fill = dangerous_fill
                    cell.font = dangerous_font

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        # Give safety padding, cap width to avoid insanely long fields
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 50)

    # Save to dynamic BytesIO stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Add dynamic date and time to the filename
    current_time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"cybershield_logs_{current_time_str}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Admin User Details Endpoint ──────────────────────────────────────────────────

@router.get("/user/details", response_model=UserDetailsOut, dependencies=[Depends(verify_admin)])
@limiter.limit("60/minute")
async def get_user_details(request: Request, email: str = Query(...), db: Session = Depends(get_db)):
    """Fetch complete user profile details & scan history combining AdminScanLog, ScanLog, and Supabase."""
    import httpx

    SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    clean_email = email.strip().lower()
    seen_inputs = set()
    history_records = []
    by_category = {}
    threats = 0
    latest_name = None

    # 1. AdminScanLog
    admin_logs = (
        db.query(AdminScanLog)
        .filter(func.lower(AdminScanLog.user_email) == clean_email)
        .order_by(AdminScanLog.created_at.desc())
        .all()
    )
    for l in admin_logs:
        if not latest_name and l.user_name:
            latest_name = l.user_name
        feat = (l.scan_type or "unknown").replace("_scan", "").upper()
        inp_snippet = (l.scan_input or "")[:40].strip().lower()
        dt_minute = _make_tz_aware(l.created_at).strftime("%Y-%m-%d %H:%M")
        key = f"{feat}-{inp_snippet}-{dt_minute}"
        if key not in seen_inputs:
            seen_inputs.add(key)
            by_category[feat] = by_category.get(feat, 0) + 1
            if (l.result or "").upper() in ["DANGEROUS", "SUSPICIOUS"]:
                threats += 1
            history_records.append(
                ScanLogOut(
                    id=l.id,
                    feature=feat,
                    input_data=l.scan_input or "",
                    verdict=l.result or "SAFE",
                    confidence=l.confidence or 0.0,
                    explanation=l.analysis or "",
                    tips=[],
                    raw={},
                    scanned_at=_make_tz_aware(l.created_at),
                    user_name=l.user_name or "User",
                    user_email=l.user_email,
                )
            )

    # 2. ScanLog
    scan_logs = (
        db.query(ScanLog)
        .filter(func.lower(ScanLog.user_email) == clean_email)
        .order_by(ScanLog.scanned_at.desc())
        .all()
    )
    for l in scan_logs:
        if not latest_name and l.user_name:
            latest_name = l.user_name
        feat = (l.feature or "unknown").replace("_scan", "").upper()
        inp_snippet = (l.input_data or "")[:40].strip().lower()
        dt_minute = _make_tz_aware(l.scanned_at).strftime("%Y-%m-%d %H:%M")
        key = f"{feat}-{inp_snippet}-{dt_minute}"
        if key not in seen_inputs:
            seen_inputs.add(key)
            by_category[feat] = by_category.get(feat, 0) + 1
            if (l.verdict or "").upper() in ["DANGEROUS", "SUSPICIOUS"]:
                threats += 1
            history_records.append(
                ScanLogOut(
                    id=l.id + 10000,
                    feature=feat,
                    input_data=l.input_data or "",
                    verdict=l.verdict or "SAFE",
                    confidence=l.confidence or 0.0,
                    explanation=l.explanation or "",
                    tips=l.tips or [],
                    raw=l.raw or {},
                    scanned_at=_make_tz_aware(l.scanned_at),
                    user_name=l.user_name or "User",
                    user_email=l.user_email,
                )
            )

    # 3. Supabase scan_logs
    if SUPABASE_URL and SERVICE_KEY:
        try:
            headers = {"apikey": SERVICE_KEY, "Authorization": f"Bearer {SERVICE_KEY}"}
            async with httpx.AsyncClient(timeout=5.0) as client:
                res = await client.get(f"{SUPABASE_URL}/rest/v1/scan_logs?select=*&order=scanned_at.desc&limit=200", headers=headers)
                if res.status_code == 200:
                    sb_logs = res.json()
                    for sb in sb_logs:
                        raw_meta = sb.get("raw") or {}
                        uemail = (raw_meta.get("user_email") or sb.get("user_email") or "").strip().lower()
                        if uemail == clean_email:
                            uname = raw_meta.get("user_name") or sb.get("user_name") or "User"
                            if not latest_name and uname:
                                latest_name = uname
                            sb_id = sb.get("id")
                            feat = (sb.get("feature") or "unknown").replace("_scan", "").upper()
                            inp_snippet = (sb.get("input_data") or "")[:40].strip().lower()
                            dt_obj = _make_tz_aware(sb.get("scanned_at"))
                            dt_minute = dt_obj.strftime("%Y-%m-%d %H:%M")
                            key = f"{feat}-{inp_snippet}-{dt_minute}"
                            if key not in seen_inputs:
                                seen_inputs.add(key)
                                by_category[feat] = by_category.get(feat, 0) + 1
                                v_res = (sb.get("verdict") or "SAFE").upper()
                                if v_res in ["DANGEROUS", "SUSPICIOUS"]:
                                    threats += 1
                                history_records.append(
                                    ScanLogOut(
                                        id=int(sb_id) + 50000 if isinstance(sb_id, int) else 50000,
                                        feature=feat,
                                        input_data=sb.get("input_data") or "",
                                        verdict=v_res,
                                        confidence=sb.get("confidence") or 0.0,
                                        explanation=sb.get("explanation") or "",
                                        tips=sb.get("tips") or [],
                                        raw=raw_meta,
                                        scanned_at=dt_obj,
                                        user_name=uname,
                                        user_email=clean_email,
                                    )
                                )
        except Exception as e:
            print(f"Error fetching user details from Supabase: {e}")

    # Sort history records
    history_records.sort(key=lambda x: _make_tz_aware(x.scanned_at), reverse=True)

    total_scans = len(history_records)
    last_scanned = history_records[0].scanned_at if history_records else None
    first_scanned = history_records[-1].scanned_at if history_records else None
    safe_rate = round(((total_scans - threats) / total_scans * 100), 1) if total_scans > 0 else 100.0

    return UserDetailsOut(
        user_name=latest_name or clean_email.split("@")[0].title(),
        user_email=clean_email,
        total_scans=total_scans,
        threats=threats,
        safe_rate=safe_rate,
        last_scanned_at=last_scanned,
        created_at=first_scanned,
        by_category=by_category,
        scan_history=history_records,
    )


# ── Admin Monitor Endpoints (admin_scan_logs table) ────────────────────────────

def _make_tz_aware(dt):
    if not dt:
        return datetime.now(timezone.utc)
    if isinstance(dt, str):
        try:
            dt = datetime.fromisoformat(dt.replace("Z", "+00:00"))
        except Exception:
            return datetime.now(timezone.utc)
    if getattr(dt, "tzinfo", None) is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt

@router.get("/monitor/scans", response_model=List[AdminMonitorScan], dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_monitor_scans(
    request: Request,
    page: int = Query(1, ge=1),
    per_page: int = Query(50, ge=1, le=200),
    scan_type: Optional[str] = None,
    result: Optional[str] = None,
    search: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sort: Optional[str] = "newest",
    db: Session = Depends(get_db),
):
    """Paginated list of scans combining local AdminScanLog, local ScanLog, and Supabase scan_logs."""
    import httpx

    SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    all_records = []
    seen_ids = set()

    # 1. Fetch from local AdminScanLog
    q1 = db.query(AdminScanLog)
    if scan_type and scan_type != "ALL":
        q1 = q1.filter(AdminScanLog.scan_type == scan_type)
    if result and result != "ALL":
        q1 = q1.filter(AdminScanLog.result == result)
    if search:
        pattern = f"%{search.strip().lower()}%"
        q1 = q1.filter(
            func.lower(AdminScanLog.user_email).like(pattern) |
            func.lower(AdminScanLog.user_name).like(pattern) |
            func.lower(AdminScanLog.scan_input).like(pattern)
        )

    for r in q1.order_by(AdminScanLog.created_at.desc()).all():
        feat_clean = (r.scan_type or "unknown").replace("_scan", "").upper()
        inp_snippet = (r.scan_input or "")[:40].strip().lower()
        dt_minute = _make_tz_aware(r.created_at).strftime("%Y-%m-%d %H:%M")
        key = f"{r.user_email or 'anon'}-{feat_clean}-{inp_snippet}-{dt_minute}"
        if key not in seen_ids:
            seen_ids.add(key)
            all_records.append(AdminMonitorScan(
                id=r.id,
                scan_id=r.scan_id or f"scan-{r.id}",
                user_id=r.user_id or "usr-001",
                user_name=r.user_name or "User",
                user_email=r.user_email or "user@cybershield.local",
                scan_type=feat_clean,
                scan_input=r.scan_input or "",
                result=r.result or "SAFE",
                confidence=r.confidence or 0.0,
                analysis=r.analysis or "Analysis complete.",
                created_at=_make_tz_aware(r.created_at)
            ))

    # 2. Fetch from local ScanLog
    q2 = db.query(ScanLog)
    if scan_type and scan_type != "ALL":
        feat_target = scan_type.lower() + "_scan" if not scan_type.endswith("_scan") else scan_type.lower()
        q2 = q2.filter(ScanLog.feature == feat_target)
    if result and result != "ALL":
        q2 = q2.filter(ScanLog.verdict == result)
    if search:
        pattern = f"%{search.strip().lower()}%"
        q2 = q2.filter(
            func.lower(ScanLog.user_email).like(pattern) |
            func.lower(ScanLog.user_name).like(pattern) |
            func.lower(ScanLog.input_data).like(pattern)
        )

    for s in q2.order_by(ScanLog.scanned_at.desc()).all():
        feat_clean = (s.feature or "unknown").replace("_scan", "").upper()
        inp_snippet = (s.input_data or "")[:40].strip().lower()
        dt_minute = _make_tz_aware(s.scanned_at).strftime("%Y-%m-%d %H:%M")
        key = f"{s.user_email or 'anon'}-{feat_clean}-{inp_snippet}-{dt_minute}"
        if key not in seen_ids:
            seen_ids.add(key)
            all_records.append(AdminMonitorScan(
                id=s.id + 10000,
                scan_id=f"scan-{s.id}",
                user_id="usr-001",
                user_name=s.user_name or "User",
                user_email=s.user_email or "user@cybershield.local",
                scan_type=feat_clean,
                scan_input=s.input_data or "",
                result=s.verdict or "SAFE",
                confidence=s.confidence or 0.0,
                analysis=s.explanation or "Analysis complete.",
                created_at=_make_tz_aware(s.scanned_at)
            ))

    # 3. Fetch from Supabase REST API
    if SUPABASE_URL and SERVICE_KEY:
        try:
            headers = {
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
            }
            async with httpx.AsyncClient(timeout=5.0) as client:
                res = await client.get(f"{SUPABASE_URL}/rest/v1/scan_logs?select=*&order=scanned_at.desc&limit=100", headers=headers)
                if res.status_code == 200:
                    sb_logs = res.json()
                    for sb in sb_logs:
                        raw_meta = sb.get("raw") or {}
                        uname = raw_meta.get("user_name") or sb.get("user_name") or "User"
                        uemail = raw_meta.get("user_email") or sb.get("user_email") or "user@cybershield.local"
                        feat_clean = (sb.get("feature") or "unknown").replace("_scan", "").upper()
                        scanned_at_val = sb.get("scanned_at")
                        dt_obj = _make_tz_aware(scanned_at_val)
                        inp_snippet = (sb.get("input_data") or "")[:40].strip().lower()
                        dt_minute = dt_obj.strftime("%Y-%m-%d %H:%M")
                        key = f"{uemail}-{feat_clean}-{inp_snippet}-{dt_minute}"
                        if key not in seen_ids:
                            seen_ids.add(key)
                            sb_id = sb.get("id")
                            all_records.append(AdminMonitorScan(
                                id=int(sb_id) + 50000 if isinstance(sb_id, int) else 50000,
                                scan_id=f"sb-{sb_id}",
                                user_id=sb.get("user_id") or "usr-001",
                                user_name=uname,
                                user_email=uemail,
                                scan_type=feat_clean,
                                scan_input=sb.get("input_data") or "",
                                result=sb.get("verdict") or "SAFE",
                                confidence=sb.get("confidence") or 0.0,
                                analysis=sb.get("explanation") or "Analysis complete.",
                                created_at=dt_obj
                            ))
        except Exception as e:
            print(f"Error fetching Supabase scan_logs: {e}")

    # Sort merged records safely with timezone-aware datetimes
    if sort == "oldest":
        all_records.sort(key=lambda x: _make_tz_aware(x.created_at))
    else:
        all_records.sort(key=lambda x: _make_tz_aware(x.created_at), reverse=True)

    # Paginate
    start = (page - 1) * per_page
    end = start + per_page
    return all_records[start:end]


@router.get("/monitor/stats", dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_monitor_stats(request: Request, db: Session = Depends(get_db)):
    """Live stats computed from admin_scan_logs."""
    total = db.query(func.count(AdminScanLog.id)).scalar() or 0
    threats = db.query(func.count(AdminScanLog.id)).filter(
        AdminScanLog.result.in_(["DANGEROUS", "SUSPICIOUS"])
    ).scalar() or 0
    safe_count = total - threats
    safe_rate = round((safe_count / total * 100), 1) if total > 0 else 100.0

    total_users = db.query(func.count(func.distinct(AdminScanLog.user_email))).scalar() or 0

    by_type = {
        row[0]: row[1]
        for row in db.query(AdminScanLog.scan_type, func.count(AdminScanLog.id))
        .group_by(AdminScanLog.scan_type).all()
    }
    by_result = {
        row[0]: row[1]
        for row in db.query(AdminScanLog.result, func.count(AdminScanLog.id))
        .group_by(AdminScanLog.result).all()
    }

    today = datetime.now(timezone.utc).date()
    scans_today = db.query(func.count(AdminScanLog.id)).filter(
        func.date(AdminScanLog.created_at) == today
    ).scalar() or 0

    daily_counts = []
    for i in range(6, -1, -1):
        day = today - timedelta(days=i)
        count = db.query(func.count(AdminScanLog.id)).filter(
            func.date(AdminScanLog.created_at) == day
        ).scalar() or 0
        daily_counts.append({"date": day.isoformat(), "count": count, "label": day.strftime("%a")})

    return {
        "total_scans": total,
        "threat_count": threats,
        "safe_count": safe_count,
        "safe_rate": safe_rate,
        "total_users": total_users,
        "scans_today": scans_today,
        "by_type": by_type,
        "by_result": by_result,
        "daily_counts": daily_counts,
    }


@router.get("/monitor/users", dependencies=[Depends(verify_admin)])
@limiter.limit("120/minute")
async def get_monitor_users(request: Request, db: Session = Depends(get_db)):
    """Per-user scan summaries from admin_scan_logs + ScanLog + Supabase Auth & public.profiles."""
    import httpx

    SUPABASE_URL = os.getenv("SUPABASE_URL", "").rstrip("/")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    # 1. Fetch scan summaries from admin_scan_logs DB table
    rows = (
        db.query(
            AdminScanLog.user_email,
            AdminScanLog.user_id,
            func.count(AdminScanLog.id).label("total_scans"),
            func.sum(
                case((AdminScanLog.result.in_(["DANGEROUS", "SUSPICIOUS"]), 1), else_=0)
            ).label("threats"),
            func.max(AdminScanLog.created_at).label("last_scanned_at"),
        )
        .group_by(AdminScanLog.user_email)
        .order_by(func.max(AdminScanLog.created_at).desc())
        .all()
    )

    user_map = {}
    for r in rows:
        email = (r.user_email or "").strip().lower()
        if not email:
            continue
        latest = (
            db.query(AdminScanLog.user_name)
            .filter(func.lower(AdminScanLog.user_email) == email)
            .order_by(AdminScanLog.created_at.desc())
            .first()
        )
        total = r.total_scans or 0
        threats = int(r.threats or 0)
        safe_rate = round(((total - threats) / total * 100), 1) if total > 0 else 100.0
        user_map[email] = {
            "user_name": latest[0] if latest and latest[0] else "User",
            "user_email": r.user_email,
            "user_id": r.user_id,
            "total_scans": total,
            "threats": threats,
            "safe_rate": safe_rate,
            "last_scanned_at": r.last_scanned_at.isoformat() if hasattr(r.last_scanned_at, "isoformat") else str(r.last_scanned_at),
        }

    # 1b. Merge scan summaries from ScanLog DB table
    scan_rows = (
        db.query(
            ScanLog.user_email,
            func.count(ScanLog.id).label("total_scans"),
            func.sum(case((ScanLog.verdict.in_(["DANGEROUS", "SUSPICIOUS"]), 1), else_=0)).label("threats"),
            func.max(ScanLog.scanned_at).label("last_scanned_at"),
        )
        .group_by(ScanLog.user_email)
        .all()
    )

    for r in scan_rows:
        email = (r.user_email or "").strip().lower()
        if not email:
            continue
        latest = db.query(ScanLog.user_name).filter(func.lower(ScanLog.user_email) == email).order_by(ScanLog.scanned_at.desc()).first()
        total = r.total_scans or 0
        threats = int(r.threats or 0)
        if email not in user_map:
            safe_rate = round(((total - threats) / total * 100), 1) if total > 0 else 100.0
            user_map[email] = {
                "user_name": latest[0] if latest and latest[0] else "User",
                "user_email": r.user_email,
                "user_id": "usr-001",
                "total_scans": total,
                "threats": threats,
                "safe_rate": safe_rate,
                "last_scanned_at": r.last_scanned_at.isoformat() if hasattr(r.last_scanned_at, "isoformat") else str(r.last_scanned_at),
            }
        else:
            if total > user_map[email]["total_scans"]:
                user_map[email]["total_scans"] = total
                user_map[email]["threats"] = max(user_map[email]["threats"], threats)
            tot_c = user_map[email]["total_scans"]
            thr_c = user_map[email]["threats"]
            user_map[email]["safe_rate"] = round(((tot_c - thr_c) / tot_c * 100), 1) if tot_c > 0 else 100.0
            if latest and latest[0] and user_map[email]["user_name"] in ["User", "Anonymous"]:
                user_map[email]["user_name"] = latest[0]

    # 2. Fetch all registered accounts from Supabase Auth Admin API & public.profiles
    if SUPABASE_URL and SERVICE_KEY:
        try:
            headers = {
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
            }
            async with httpx.AsyncClient(timeout=5.0) as client:
                res = await client.get(f"{SUPABASE_URL}/auth/v1/admin/users", headers=headers)
                if res.status_code == 200:
                    auth_users = res.json().get("users", [])
                    for u in auth_users:
                        em = (u.get("email") or "").strip().lower()
                        if not em:
                            continue
                        meta = u.get("user_metadata") or {}
                        full_name = meta.get("full_name") or u.get("email", "").split("@")[0] or "Registered User"
                        created_at = u.get("created_at")
                        if em not in user_map:
                            user_map[em] = {
                                "user_name": full_name,
                                "user_email": u.get("email"),
                                "user_id": u.get("id"),
                                "total_scans": 0,
                                "threats": 0,
                                "safe_rate": 100.0,
                                "last_scanned_at": created_at,
                            }
                        else:
                            if full_name and user_map[em]["user_name"] in ["User", "Anonymous"]:
                                user_map[em]["user_name"] = full_name

                # Also fetch from public.profiles table using service role key
                prof_res = await client.get(f"{SUPABASE_URL}/rest/v1/profiles?select=*", headers=headers)
                if prof_res.status_code == 200:
                    profiles_list = prof_res.json()
                    for p in profiles_list:
                        em = (p.get("email") or "").strip().lower()
                        if not em:
                            continue
                        full_name = p.get("full_name") or p.get("username") or em.split("@")[0]
                        created_at = p.get("created_at")
                        if em not in user_map:
                            user_map[em] = {
                                "user_name": full_name,
                                "user_email": p.get("email"),
                                "user_id": p.get("id"),
                                "total_scans": 0,
                                "threats": 0,
                                "safe_rate": 100.0,
                                "last_scanned_at": created_at,
                            }
                        else:
                            if full_name and user_map[em]["user_name"] in ["User", "Anonymous"]:
                                user_map[em]["user_name"] = full_name
        except Exception as e:
            print(f"Error fetching Supabase users/profiles: {e}")

    # Convert map to list sorted by last activity
    results = list(user_map.values())
    results.sort(key=lambda x: x.get("last_scanned_at") or "", reverse=True)
    return results



@router.delete("/users/{user_email:path}", dependencies=[Depends(verify_admin)])
@limiter.limit("60/minute")
async def delete_user_by_email(request: Request, user_email: str, db: Session = Depends(get_db)):
    """Permanently delete a user from SQLite + all Supabase tables + auth."""
    import httpx

    SUPABASE_URL = os.getenv("SUPABASE_URL", "")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    email_clean = user_email.strip().lower()
    if email_clean == "varaprasadmokharala5@gmail.com":
        raise HTTPException(status_code=403, detail="Cannot delete the Primary Admin account.")

    # ── 1. Delete from local SQLite ───────────────────────────────────────────
    deleted_scan_logs  = db.query(ScanLog).filter(
        func.lower(ScanLog.user_email) == email_clean
    ).delete(synchronize_session=False)

    deleted_admin_logs = db.query(AdminScanLog).filter(
        func.lower(AdminScanLog.user_email) == email_clean
    ).delete(synchronize_session=False)

    db.commit()

    supabase_auth_deleted = False
    tables_deleted        = []
    supabase_error        = None

    if SUPABASE_URL and SERVICE_KEY:
        headers = {
            "apikey":        SERVICE_KEY,
            "Authorization": f"Bearer {SERVICE_KEY}",
            "Content-Type":  "application/json",
            "Prefer":        "return=minimal",
        }

        async with httpx.AsyncClient(timeout=15) as client:

            # ── 2. Delete from Supabase public tables (service role bypasses RLS) ──
            rest_base = f"{SUPABASE_URL}/rest/v1"

            # Delete scan_logs by user_email
            r = await client.delete(f"{rest_base}/scan_logs",
                headers=headers, params={"user_email": f"eq.{email_clean}"})
            if r.status_code in (200, 204): tables_deleted.append("scan_logs")

            # Delete admin_scan_logs by user_email
            r = await client.delete(f"{rest_base}/admin_scan_logs",
                headers=headers, params={"user_email": f"eq.{email_clean}"})
            if r.status_code in (200, 204): tables_deleted.append("admin_scan_logs")

            # Look up profile id first (needed for login_activity)
            prof_resp = await client.get(f"{rest_base}/profiles",
                headers=headers, params={"email": f"eq.{email_clean}", "select": "id"})
            profile_id = None
            if prof_resp.status_code == 200:
                rows = prof_resp.json()
                if rows:
                    profile_id = rows[0].get("id")

            # Delete login_activity by user_id
            if profile_id:
                r = await client.delete(f"{rest_base}/login_activity",
                    headers=headers, params={"user_id": f"eq.{profile_id}"})
                if r.status_code in (200, 204): tables_deleted.append("login_activity")

            # Delete from public.users
            r = await client.delete(f"{rest_base}/users",
                headers=headers, params={"email": f"eq.{email_clean}"})
            if r.status_code in (200, 204): tables_deleted.append("users")

            # Delete from public.profiles  ← this removes user from admin portal
            r = await client.delete(f"{rest_base}/profiles",
                headers=headers, params={"email": f"eq.{email_clean}"})
            if r.status_code in (200, 204): tables_deleted.append("profiles")

            # ── 3. Delete from Supabase auth.users ───────────────────────────────
            try:
                lookup = await client.get(
                    f"{SUPABASE_URL}/auth/v1/admin/users",
                    headers=headers,
                    params={"filter": email_clean, "page": 1, "per_page": 10},
                )
                if lookup.status_code == 200:
                    users_data = lookup.json()
                    user_list  = users_data if isinstance(users_data, list) else users_data.get("users", [])
                    matched    = [u for u in user_list if (u.get("email") or "").lower() == email_clean]
                    for u in matched:
                        uid = u.get("id")
                        if uid:
                            del_resp = await client.delete(
                                f"{SUPABASE_URL}/auth/v1/admin/users/{uid}",
                                headers=headers,
                            )
                            if del_resp.status_code in (200, 204):
                                supabase_auth_deleted = True
                else:
                    supabase_error = f"Auth lookup {lookup.status_code}"
            except Exception as e:
                supabase_error = str(e)

    return {
        "success":              True,
        "message":              f"User '{email_clean}' permanently deleted.",
        "deleted_local_scans":  deleted_scan_logs + deleted_admin_logs,
        "supabase_tables":      tables_deleted,
        "supabase_auth_deleted": supabase_auth_deleted,
        "supabase_error":       supabase_error,
    }


@router.post("/scans/reset", dependencies=[Depends(verify_admin)])
@limiter.limit("10/minute")
async def reset_all_scans(request: Request, db: Session = Depends(get_db)):
    """Clear/Reset all scan history across local SQLite and Supabase."""
    import httpx

    SUPABASE_URL = os.getenv("SUPABASE_URL", "")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    # 1. Delete local SQLite scan logs
    del_admin_logs = db.query(AdminScanLog).delete(synchronize_session=False)
    del_scan_logs = db.query(ScanLog).delete(synchronize_session=False)
    db.commit()

    sb_deleted = False
    if SUPABASE_URL and SERVICE_KEY:
        try:
            headers = {
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
                "Content-Type": "application/json",
            }
            async with httpx.AsyncClient(timeout=10) as client:
                res = await client.delete(f"{SUPABASE_URL}/rest/v1/scan_logs?id=gt.0", headers=headers)
                if res.status_code in (200, 204):
                    sb_deleted = True
        except Exception as e:
            print(f"Error resetting Supabase scan_logs: {e}")

    return {
        "success": True,
        "message": "All scan logs have been reset successfully.",
        "deleted_admin_logs": del_admin_logs,
        "deleted_scan_logs": del_scan_logs,
        "supabase_reset": sb_deleted
    }


@router.post("/user/reset-scans", dependencies=[Depends(verify_admin)])
@limiter.limit("30/minute")
async def reset_user_scans(request: Request, email: str = Query(...), db: Session = Depends(get_db)):
    """Reset scan history for a specific user."""
    import httpx

    clean_email = email.strip().lower()
    SUPABASE_URL = os.getenv("SUPABASE_URL", "")
    SERVICE_KEY  = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

    del_admin = db.query(AdminScanLog).filter(func.lower(AdminScanLog.user_email) == clean_email).delete(synchronize_session=False)
    del_scans = db.query(ScanLog).filter(func.lower(ScanLog.user_email) == clean_email).delete(synchronize_session=False)
    db.commit()

    sb_deleted = False
    if SUPABASE_URL and SERVICE_KEY:
        try:
            headers = {
                "apikey": SERVICE_KEY,
                "Authorization": f"Bearer {SERVICE_KEY}",
            }
            async with httpx.AsyncClient(timeout=10) as client:
                res = await client.delete(f"{SUPABASE_URL}/rest/v1/scan_logs?user_email=eq.{clean_email}", headers=headers)
                if res.status_code in (200, 204):
                    sb_deleted = True
        except Exception as e:
            print(f"Error resetting user scans in Supabase: {e}")

    return {
        "success": True,
        "message": f"Scan history for {clean_email} reset successfully.",
        "deleted_local": del_admin + del_scans,
        "supabase_reset": sb_deleted
    }

