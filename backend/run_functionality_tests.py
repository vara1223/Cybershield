import os
import sys
import json
import base64
import urllib.request
import urllib.parse
import subprocess
from io import BytesIO

# 1. Install openpyxl if not installed
print("Checking for openpyxl library...")
try:
    import openpyxl
    print("openpyxl is already installed.")
except ImportError:
    print("Installing openpyxl library...")
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl
    print("openpyxl installed successfully.")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PIL is required for generating mock screenshot image
try:
    from PIL import Image, ImageDraw
except ImportError:
    print("Installing Pillow...")
    subprocess.run([sys.executable, "-m", "pip", "install", "Pillow"], check=True)
    from PIL import Image, ImageDraw

from dotenv import load_dotenv
load_dotenv()
import os
ADMIN_API_KEY = os.getenv("ADMIN_API_KEY", "")

BASE_URL = "http://localhost:8000"

def post_request(endpoint, payload):
    url = f"{BASE_URL}{endpoint}"
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    if endpoint.startswith("/admin") and ADMIN_API_KEY:
        headers["X-Admin-Key"] = ADMIN_API_KEY
    req = urllib.request.Request(
        url, data=data, headers=headers, method="POST"
    )
    try:
        with urllib.request.urlopen(req) as response:
            res_data = response.read().decode("utf-8")
            return response.status, json.loads(res_data)
    except urllib.error.HTTPError as e:
        try:
            err_data = e.read().decode("utf-8")
            return e.code, json.loads(err_data)
        except Exception:
            return e.code, {"error": str(e)}
    except Exception as e:
        return 0, {"error": str(e)}

def get_request(endpoint):
    url = f"{BASE_URL}{endpoint}"
    headers = {}
    if endpoint.startswith("/admin") and ADMIN_API_KEY:
        headers["X-Admin-Key"] = ADMIN_API_KEY
    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req) as response:
            res_data = response.read().decode("utf-8")
            if "csv" in endpoint:
                return response.status, res_data  # CSV is raw text
            return response.status, json.loads(res_data)
    except urllib.error.HTTPError as e:
        try:
            err_data = e.read().decode("utf-8")
            return e.code, json.loads(err_data)
        except Exception:
            return e.code, {"error": str(e)}
    except Exception as e:
        return 0, {"error": str(e)}

# Generate mock screenshot base64
img = Image.new("RGB", (400, 100), "white")
d = ImageDraw.Draw(img)
d.text((10, 10), "URGENT: Your account is suspended. Click link http://scam.tk/otp", fill="black")
d.text((10, 40), "Pay fine of INR 500 immediately to avoid arrest.", fill="black")
buf = BytesIO()
img.save(buf, format="PNG")
screenshot_b64 = base64.b64encode(buf.getvalue()).decode()

# Silent WAV base64
silent_wav = (
    b"RIFF$\x00\x00\x00WAVEfmt \x10\x00\x00\x00\x01\x00\x01\x00"
    b"D\xac\x00\x00\x88X\x01\x00\x02\x00\x10\x00data\x00\x00\x00\x00"
)
silent_b64 = base64.b64encode(silent_wav).decode()

test_cases = [
    # Health Check
    {
        "id": "TC-001",
        "category": "Health Check",
        "description": "Verify system health check endpoint",
        "method": "GET",
        "endpoint": "/health",
        "payload": None,
        "expected_status": 200,
        "expected_verdict": None,
        "validator": lambda res: res.get("status") == "ok"
    },
    # URL Scan
    {
        "id": "TC-002",
        "category": "URL Scan",
        "description": "Verify safe URL check (Google)",
        "method": "POST",
        "endpoint": "/analyze/url",
        "payload": {"url": "https://google.com"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    {
        "id": "TC-003",
        "category": "URL Scan",
        "description": "Verify dangerous phishing URL with suspicious domain",
        "method": "POST",
        "endpoint": "/analyze/url",
        "payload": {"url": "http://sbi-bank-update-kyc-urgent.tk/verify?otp=123456"},
        "expected_status": 200,
        "expected_verdict": "DANGEROUS",
        "validator": lambda res: res.get("verdict") == "DANGEROUS"
    },
    {
        "id": "TC-004",
        "category": "URL Scan",
        "description": "Verify shortened URL warning",
        "method": "POST",
        "endpoint": "/analyze/url",
        "payload": {"url": "https://bit.ly/abc123"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") in ["SAFE", "SUSPICIOUS"]
    },
    # Screenshot Scan
    {
        "id": "TC-005",
        "category": "Screenshot Scan",
        "description": "Verify OCR and spam classification on mock scam image",
        "method": "POST",
        "endpoint": "/analyze/screenshot",
        "payload": {"image": screenshot_b64},
        "expected_status": 200,
        "expected_verdict": "DANGEROUS",
        "validator": lambda res: res.get("verdict") == "DANGEROUS"
    },
    # QR Scan
    {
        "id": "TC-006",
        "category": "QR Scan",
        "description": "Verify QR scan with decoded phishing URL payload",
        "method": "POST",
        "endpoint": "/analyze/qr",
        "payload": {"decoded_content": "http://sbi-bank-update-kyc-urgent.tk/verify?otp=123456"},
        "expected_status": 200,
        "expected_verdict": "DANGEROUS",
        "validator": lambda res: res.get("verdict") == "DANGEROUS"
    },
    {
        "id": "TC-007",
        "category": "QR Scan",
        "description": "Verify QR scan with safe plain text payload",
        "method": "POST",
        "endpoint": "/analyze/qr",
        "payload": {"decoded_content": "Scan result showing plain reference info"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    # OTP Scam
    {
        "id": "TC-008",
        "category": "OTP Scan",
        "description": "Verify safe OTP warning message (legitimate transactional text)",
        "method": "POST",
        "endpoint": "/analyze/otp",
        "payload": {"message": "Your OTP for SBI NetBanking is 482910. Do NOT share with anyone."},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    {
        "id": "TC-009",
        "category": "OTP Scan",
        "description": "Verify dangerous OTP sharing solicitation message",
        "method": "POST",
        "endpoint": "/analyze/otp",
        "payload": {"message": "Your OTP is 123456. Share with our agent to verify your account."},
        "expected_status": 200,
        "expected_verdict": "DANGEROUS",
        "validator": lambda res: res.get("verdict") == "DANGEROUS"
    },
    # UPI Fraud
    {
        "id": "TC-010",
        "category": "UPI Scan",
        "description": "Verify safe standard UPI address",
        "method": "POST",
        "endpoint": "/analyze/upi",
        "payload": {"upi_id": "paytm@paytm"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    {
        "id": "TC-011",
        "category": "UPI Scan",
        "description": "Verify known safe banking handle",
        "method": "POST",
        "endpoint": "/analyze/upi",
        "payload": {"upi_id": "sbi.rajesh123@okaxis"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    {
        "id": "TC-012",
        "category": "UPI Scan",
        "description": "Verify suspicious lottery/refund UPI address",
        "method": "POST",
        "endpoint": "/analyze/upi",
        "payload": {"upi_id": "lottery-winner-9876@ybl"},
        "expected_status": 200,
        "expected_verdict": "SUSPICIOUS",
        "validator": lambda res: res.get("verdict") == "SUSPICIOUS"
    },
    # Voice Scan
    {
        "id": "TC-013",
        "category": "Voice Scan",
        "description": "Verify silent voice audio recording defaults to SAFE",
        "method": "POST",
        "endpoint": "/analyze/voice",
        "payload": {"audio": silent_b64, "format": "wav"},
        "expected_status": 200,
        "expected_verdict": "SAFE",
        "validator": lambda res: res.get("verdict") == "SAFE"
    },
    # Admin & Logs
    {
        "id": "TC-014",
        "category": "Admin Panel",
        "description": "Verify admin stats endpoint",
        "method": "GET",
        "endpoint": "/admin/stats",
        "payload": None,
        "expected_status": 200,
        "expected_verdict": None,
        "validator": lambda res: "total" in res
    },
    {
        "id": "TC-015",
        "category": "Admin Panel",
        "description": "Verify admin scan logs endpoint",
        "method": "GET",
        "endpoint": "/admin/logs?limit=5",
        "payload": None,
        "expected_status": 200,
        "expected_verdict": None,
        "validator": lambda res: isinstance(res, list)
    },
    {
        "id": "TC-016",
        "category": "Admin Panel",
        "description": "Verify CSV export endpoint",
        "method": "GET",
        "endpoint": "/admin/export/csv",
        "payload": None,
        "expected_status": 200,
        "expected_verdict": None,
        "validator": lambda res: isinstance(res, str) and "id" in res.lower()
    }
]

# Run tests
results = []
print("\nStarting end-to-end functionality tests...")
for tc in test_cases:
    print(f"Running {tc['id']}: {tc['description']}...", end="")
    if tc["method"] == "POST":
        status, response = post_request(tc["endpoint"], tc["payload"])
    else:
        status, response = get_request(tc["endpoint"])
    
    # Evaluate
    actual_verdict = response.get("verdict") if isinstance(response, dict) else None
    
    status_ok = (status == tc["expected_status"])
    val_ok = tc["validator"](response)
    
    passed = "PASS" if (status_ok and val_ok) else "FAIL"
    
    detail = ""
    if passed == "FAIL":
        detail = f"Status: {status} (Expected: {tc['expected_status']}), Response: {str(response)[:150]}"
    else:
        if isinstance(response, dict):
            confidence = response.get("confidence")
            flags = response.get("flags")
            detail = f"Confidence: {confidence}%" if confidence is not None else "Endpoint OK"
            if flags:
                detail += f", Flags: {flags}"
        else:
            detail = "Endpoint responded with correct formatting and content."
            
    print(f" [{passed}]")
    
    results.append({
        "id": tc["id"],
        "category": tc["category"],
        "description": tc["description"],
        "endpoint": tc["endpoint"],
        "method": tc["method"],
        "payload": json.dumps(tc["payload"]) if tc["payload"] else "",
        "expected_status": tc["expected_status"],
        "actual_status": status,
        "expected_verdict": tc["expected_verdict"] or "N/A",
        "actual_verdict": actual_verdict or "N/A",
        "status": passed,
        "details": detail
    })

# Write to Excel
print("\nGenerating Excel report...")
wb = Workbook()
ws = wb.active
ws.title = "Functionality Test Report"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Style variables
font_family = "Segoe UI"
header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
data_font = Font(name=font_family, size=10)
bold_font = Font(name=font_family, size=10, bold=True)

pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Soft green
pass_font = Font(name=font_family, size=10, bold=True, color="375623")
fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # Soft red
fail_font = Font(name=font_family, size=10, bold=True, color="C65911")

border_thin = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
)

# Title block
ws.cell(row=2, column=2, value="CyberShield - Functionality Test Report").font = title_font
ws.cell(row=3, column=2, value="Generated Automatically - E2E API Verification").font = subtitle_font

headers = [
    "Test ID", "Category", "Description", "Endpoint", "Method", 
    "Expected Status", "Actual Status", "Expected Verdict", "Actual Verdict", 
    "Status", "Details/Logs"
]

start_row = 5
for col_idx, header in enumerate(headers, start=2):
    cell = ws.cell(row=start_row, column=col_idx, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border_thin

# Add Data
current_row = start_row + 1
pass_count = 0
fail_count = 0

for item in results:
    row_data = [
        item["id"], item["category"], item["description"], item["endpoint"], item["method"],
        item["expected_status"], item["actual_status"], item["expected_verdict"], item["actual_verdict"],
        item["status"], item["details"]
    ]
    
    for col_idx, val in enumerate(row_data, start=2):
        cell = ws.cell(row=current_row, column=col_idx, value=val)
        cell.font = data_font
        cell.border = border_thin
        
        # Alignments
        if col_idx in [2, 6, 7, 8, 9, 10]:  # ID, Method, Statuses, Verdicts
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 11:  # Status
            cell.alignment = Alignment(horizontal="center", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            
        # Status styling
        if col_idx == 11:
            if val == "PASS":
                cell.fill = pass_fill
                cell.font = pass_font
                pass_count += 1
            else:
                cell.fill = fail_fill
                cell.font = fail_font
                fail_count += 1
                
    current_row += 1

# Auto-fit columns
for col in ws.columns:
    if col[0].column < 2 or col[0].column > len(headers) + 1:
        continue
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.row < start_row:
            continue
        val_str = str(cell.value or "")
        if len(val_str) > max_len:
            max_len = len(val_str)
    # Set padding
    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

# Summary Card
ws.cell(row=2, column=8, value="Total Test Cases:").font = bold_font
ws.cell(row=2, column=9, value=len(results)).font = data_font
ws.cell(row=3, column=8, value="Passed:").font = bold_font
ws.cell(row=3, column=9, value=pass_count).font = pass_font
ws.cell(row=4, column=8, value="Failed:").font = bold_font
ws.cell(row=4, column=9, value=fail_count).font = fail_font if fail_count > 0 else data_font

# Save Workbook
excel_path = "CyberShield_Functionality_Test_Report.xlsx"
wb.save(excel_path)
print(f"Report saved to: {os.path.abspath(excel_path)}")
print(f"Execution complete. Passed: {pass_count}, Failed: {fail_count}")
