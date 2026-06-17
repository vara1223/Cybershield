from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import time
import os

def run_real_selenium_tests():
    print("Initializing headless Chrome WebDriver...")
    options = webdriver.ChromeOptions()
    options.add_argument("--headless")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(options=options)
    wait = WebDriverWait(driver, 10)
    
    results = []
    
    # Help helper function to log results
    def log_result(tc_id, category, desc, locator, status, details):
        results.append({
            "id": tc_id,
            "category": category,
            "description": desc,
            "locator": locator,
            "status": status,
            "details": details,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
        print(f"[{status}] {tc_id}: {desc}")

    try:
        # TC-WEB-001: Load Home Page
        tc_id = "TC-WEB-001"
        category = "Page Load"
        desc = "Navigate to http://localhost:8081 and verify page title is 'CyberShield' or 'Login'"
        try:
            driver.get("http://localhost:8081")
            time.sleep(5)  # Wait for Expo package bundle load
            title = driver.title
            if "CyberShield" in title or "Login" in title:
                log_result(tc_id, category, desc, "driver.title", "PASS", f"Successfully loaded home page. Title: '{title}'")
            else:
                log_result(tc_id, category, desc, "driver.title", "FAIL", f"Page title '{title}' did not match expected values")
        except Exception as e:
            log_result(tc_id, category, desc, "driver.get", "FAIL", f"Navigation failed: {str(e)[:100]}")

        # TC-WEB-002: Enter Admin Credentials
        tc_id = "TC-WEB-002"
        category = "Navigation"
        desc = "Enter admin credentials and verify Passkey modal loads"
        try:
            email_input = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[@placeholder='Email']")
            ))
            email_input.clear()
            email_input.send_keys("admin@cybershield.com")
            
            pass_input = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[@placeholder='Password']")
            ))
            pass_input.clear()
            pass_input.send_keys("admin123")
            
            login_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[text()='Log In']")
            ))
            login_btn.click()
            time.sleep(2)
            
            # Check if Passkey Title is visible
            pin_title = wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), 'Enter 4-digit Passkey')]")
            ))
            log_result(tc_id, category, desc, "admin credentials login", "PASS", "Redirection to Passkey entry screen succeeded.")
        except Exception as e:
            log_result(tc_id, category, desc, "admin credentials login", "FAIL", f"Failed to navigate: {str(e)[:100]}")

        # TC-WEB-003: Enter PIN
        tc_id = "TC-WEB-003"
        category = "Authentication"
        desc = "Enter PIN '1234' on key pad and verify dashboard authorization"
        try:
            # Locate num keys 1, 2, 3, 4 and click them
            for digit in ['1', '2', '3', '4']:
                key = wait.until(EC.element_to_be_clickable(
                    (By.XPATH, f"//*[text()='{digit}']")
                ))
                key.click()
                time.sleep(0.3)
            
            # Wait for admin authenticated dashboard view to render
            threat_overview = wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(text(), 'Admin Panel')]")
            ))
            log_result(tc_id, category, desc, "numpad keys 1-4", "PASS", "Dashboard authenticated successfully with PIN '1234'.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            try:
                print("=== BROWSER CONSOLE LOGS ===")
                for entry in driver.get_log('browser'):
                    print(entry)
                print("============================")
            except Exception as le:
                print("Failed to get browser logs:", le)
            try:
                print("=== PIN PAGE DOM SOURCE ===")
                print(driver.page_source)
                print("===========================")
            except:
                pass
            log_result(tc_id, category, desc, "numpad keys 1-4", "FAIL", f"Failed to authenticate: {str(e)[:100]}")

        # TC-WEB-004: Dashboard Verification
        tc_id = "TC-WEB-004"
        category = "Dashboard UI"
        desc = "Verify presence of stats cards (Total scans, Threats, Safe rate)"
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Total Scans')]")))
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Threats Blocked')]")))
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Safe Rate')]")))
            log_result(tc_id, category, desc, "xpath(Total scans / Threats)", "PASS", "All standard stats cards are successfully displayed in dashboard.")
        except Exception as e:
            log_result(tc_id, category, desc, "xpath(Total scans / Threats)", "FAIL", f"Failed to verify cards: {str(e)[:100]}")

        # TC-WEB-005: Chart Verification
        tc_id = "TC-WEB-005"
        category = "Dashboard UI"
        desc = "Verify weekly activity chart container is present"
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'WEEKLY THREAT VOLUME')]")))
            log_result(tc_id, category, desc, "xpath(WEEKLY THREAT VOLUME)", "PASS", "Weekly activity chart container verified.")
        except Exception as e:
            log_result(tc_id, category, desc, "xpath(WEEKLY THREAT VOLUME)", "FAIL", f"Chart not found: {str(e)[:100]}")

        # TC-WEB-006: Lock Dashboard / Logout
        tc_id = "TC-WEB-006"
        category = "Navigation"
        desc = "Click 'Lock panel' button in Admin and verify redirection back to PIN screen"
        try:
            # Click on Diagnostics tab to show the lock button
            diag_tab = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), 'Diagnostics')]")
            ))
            diag_tab.click()
            time.sleep(1)

            lock_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//*[contains(text(), 'Lock Admin Console')]")
            ))
            # Try normal click, fallback to JS executor click
            try:
                lock_btn.click()
            except:
                driver.execute_script("arguments[0].click();", lock_btn)
                
            time.sleep(1)
            # Verify we are back on the Passkey entry screen
            wait.until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(), 'Enter 4-digit Passkey')]")))
            log_result(tc_id, category, desc, "xpath(//*[contains(text(), 'Lock Admin Console')])", "PASS", "Dashboard successfully locked and redirected to Passkey screen.")
        except Exception as e:
            log_result(tc_id, category, desc, "xpath(//*[contains(text(), 'Lock Admin Console')])", "FAIL", f"Failed to lock dashboard: {str(e)[:100]}")

    except Exception as e:
        print(f"Critical execution error: {e}")
    finally:
        driver.quit()
        
    generate_excel_report(results)

def generate_excel_report(results):
    print("Generating Selenium Excel report...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Selenium Web Test Report"
    
    ws.views.sheetView[0].showGridLines = True
    
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="31859C", end_color="31859C", fill_type="solid") # Teal Theme
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    title_font = Font(name=font_family, size=16, bold=True, color="31859C")
    subtitle_font = Font(name=font_family, size=10, italic=True, color="595959")
    data_font = Font(name=font_family, size=10)
    bold_font = Font(name=font_family, size=10, bold=True)
    
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    pass_font = Font(name=font_family, size=10, bold=True, color="375623")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    fail_font = Font(name=font_family, size=10, bold=True, color="C65911")
    
    border_thin = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )
    
    ws.cell(row=2, column=2, value="CyberShield - Web Page Selenium Report").font = title_font
    ws.cell(row=3, column=2, value="Executed on Headless Chrome - Live Web UI Path Verification").font = subtitle_font
    
    headers = ["Test ID", "Category", "Description", "Locator / Variable", "Status", "Execution Details", "Timestamp"]
    
    start_row = 5
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_thin
        
    current_row = start_row + 1
    pass_count = 0
    fail_count = 0
    
    for item in results:
        row_data = [
            item["id"], item["category"], item["description"], item["locator"], 
            item["status"], item["details"], item["timestamp"]
        ]
        
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = border_thin
            
            if col_idx in [2, 3, 6, 8]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            if col_idx == 6:
                if val == "PASS":
                    cell.fill = pass_fill
                    cell.font = pass_font
                    pass_count += 1
                else:
                    cell.fill = fail_fill
                    cell.font = fail_font
                    fail_count += 1
                    
        current_row += 1
        
    for col in ws.columns:
        if col[0].column < 2 or col[0].column > len(headers) + 1:
            continue
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < start_row:
                continue
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)
        
    ws.cell(row=2, column=7, value="Total Web Tests:").font = bold_font
    ws.cell(row=2, column=8, value=len(results)).font = data_font
    ws.cell(row=3, column=7, value="Passed:").font = bold_font
    ws.cell(row=3, column=8, value=pass_count).font = pass_font
    ws.cell(row=4, column=7, value="Failed:").font = bold_font
    ws.cell(row=4, column=8, value=fail_count).font = fail_font if fail_count > 0 else data_font
    
    excel_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "CyberShield_Selenium_Test_Report.xlsx"))
    wb.save(excel_path)
    print(f"Report generated successfully at: {excel_path}")

if __name__ == '__main__':
    run_real_selenium_tests()
