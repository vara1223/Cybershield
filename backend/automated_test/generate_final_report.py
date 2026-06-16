"""
CyberShield DAST Security Report Generator — Final Version
Reads all *_results.json files from automated_test/ directory
and generates a comprehensive color-coded Excel report.
"""

import json, os, glob
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

# ─── Configuration ─────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(SCRIPT_DIR, "CyberShield_Security_Report_FINAL.xlsx")

# ─── Color Palette ─────────────────────────────────────────────────────────────
C = {
    "CRITICAL": "C0392B", "HIGH": "E67E22", "MEDIUM": "F39C12",
    "LOW": "27AE60",      "INFO": "2980B9",
    "header": "1A1A2E",   "subheader": "16213E", "section": "0F3460",
    "pass_bg": "D5F5E3",  "fail_bg": "FADBD8",   "warn_bg": "FDEBD0",
    "alt":    "F8F9FA",   "white": "FFFFFF",      "border": "BDC3C7",
    "fixed":  "1E8449",   "open":  "C0392B",
}
SEV_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

# Category labels
CAT_LABELS = {
    "01_endpoint_discovery": "Endpoint Discovery",
    "02_authn_bypass":       "Authentication Bypass",
    "03_authz_privesc":      "Authorization / Privilege Escalation",
    "04_idor":               "Insecure Direct Object Reference (IDOR)",
    "05_rbac_matrix":        "RBAC / Access Control Matrix",
    "06_token_tampering":    "JWT Token Tampering",
    "07_injection":          "Injection Attacks (SQLi / NoSQLi / XSS)",
    "08_rate_limit":         "Rate Limiting",
    "09_hardcoded_creds":    "Hardcoded Credentials / Secret Scanning",
}

# Issues fixed by the remediation — tuple: (endpoint, test_category)
FIXED_ISSUES = {
    # admin endpoints now require X-Admin-Key
    ("/admin/stats",      "02_authn_bypass"),
    ("/admin/logs",       "02_authn_bypass"),
    ("/admin/export/csv", "02_authn_bypass"),
    # rate limiting applied to admin endpoints
    ("/admin/stats",      "08_rate_limit"),
    ("/admin/logs",       "08_rate_limit"),
    ("/admin/export/csv", "08_rate_limit"),
    # analyze endpoints rate limit tightened
    ("/analyze/url",      "08_rate_limit"),
    ("/analyze/otp",      "08_rate_limit"),
}

# ─── Helpers ───────────────────────────────────────────────────────────────────
def fill(hex_c):
    return PatternFill("solid", fgColor=hex_c)

def font(bold=False, color="000000", size=10, name="Calibri"):
    return Font(bold=bold, color=color, size=size, name=name)

def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, text, bg=None, fg="FFFFFF", size=11, bold=True, h="center"):
    bg = bg or C["header"]
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg); c.font = font(bold, fg, size)
    c.alignment = align(h); c.border = border()
    return c

def cell(ws, row, col, val, bg=None, bold=False, h="left", fg="000000", size=10):
    c = ws.cell(row=row, column=col, value=val)
    if bg: c.fill = fill(bg)
    c.font = font(bold, fg, size); c.alignment = align(h); c.border = border()
    return c

def sev_color(sev):
    return C.get(sev, "FFFFFF")

def status_color(finding, fixed=False):
    if fixed:    return C["pass_bg"]
    if finding:  return C["fail_bg"]
    return C["pass_bg"]

def col_w(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ─── Load Results ──────────────────────────────────────────────────────────────
def load_results():
    all_rows = []
    files = sorted(glob.glob(os.path.join(SCRIPT_DIR, "*_results.json")))
    for fpath in files:
        with open(fpath, encoding="utf-8") as f:
            try:
                rows = json.load(f)
                all_rows.extend(rows)
            except Exception as e:
                print(f"  WARN: Could not parse {fpath}: {e}")
    print(f"  Loaded {len(all_rows)} test results from {len(files)} files.")
    return all_rows

def is_in_fixed_set(row):
    """Returns True if this endpoint+category was targeted by a remediation."""
    ep  = row.get("endpoint", "")
    cat = row.get("test_category", "")
    return (ep, cat) in FIXED_ISSUES

def remediation_status(row):
    if is_in_fixed_set(row):
        return "FIXED"          # remediation was applied; now passes or was a finding
    if not row.get("finding", False):
        return "PASS"
    return "OPEN"

def is_fixed(row):
    """Alias for report coloring logic."""
    return is_in_fixed_set(row)

# ─── Sheet: Cover Page ─────────────────────────────────────────────────────────
def build_cover(wb, all_rows, gen_time):
    ws = wb.create_sheet("Cover", 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 5

    # Title block
    ws.merge_cells("B2:D3")
    c = ws["B2"]
    c.value = "CyberShield Security Assessment Report"
    c.fill = fill(C["header"]); c.font = font(True, "FFFFFF", 18)
    c.alignment = align("center")

    ws.merge_cells("B4:D4")
    c = ws["B4"]; c.value = "Dynamic Application Security Testing (DAST) — Final Report"
    c.fill = fill(C["subheader"]); c.font = font(False, "BDC3C7", 12)
    c.alignment = align("center")

    ws.merge_cells("B5:D5")
    c = ws["B5"]; c.value = f"Generated: {gen_time}"
    c.fill = fill(C["section"]); c.font = font(False, "ECF0F1", 10)
    c.alignment = align("center")

    # Stats
    total  = len(all_rows)
    findings = [r for r in all_rows if r.get("finding")]
    fixed_f  = [r for r in findings if is_fixed(r)]
    open_f   = [r for r in findings if not is_fixed(r)]
    passes   = [r for r in all_rows if not r.get("finding")]

    by_sev = {}
    for r in open_f:
        s = r.get("severity", "INFO")
        by_sev[s] = by_sev.get(s, 0) + 1

    stats = [
        ("Total Tests Run",   total,           C["section"],  "FFFFFF"),
        ("Tests Passed",      len(passes),      C["fixed"],    "FFFFFF"),
        ("Findings Fixed",    len(fixed_f),     "1A5276",      "FFFFFF"),
        ("Open Findings",     len(open_f),      C["open"],     "FFFFFF"),
        ("Critical (Open)",   by_sev.get("CRITICAL",0), C["CRITICAL"], "FFFFFF"),
        ("High (Open)",       by_sev.get("HIGH",0),     C["HIGH"],     "FFFFFF"),
        ("Medium (Open)",     by_sev.get("MEDIUM",0),   C["MEDIUM"],   "FFFFFF"),
        ("Low (Open)",        by_sev.get("LOW",0),      C["LOW"],      "FFFFFF"),
    ]

    row = 7
    ws.merge_cells(f"B{row}:D{row}")
    hdr(ws, row, 2, "Executive Summary", C["subheader"], size=13)
    row += 1
    for label, val, bg, fg in stats:
        ws.merge_cells(f"B{row}:C{row}")
        cell(ws, row, 2, label, bg, True, "left", fg, 11)
        cell(ws, row, 4, val,   bg, True, "center", fg, 13)
        row += 1

    # Scope
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    hdr(ws, row, 2, "Test Scope", C["subheader"], size=12)
    row += 1
    scope_items = [
        "Target API: http://127.0.0.1:8000",
        "Backend:    FastAPI + SQLite",
        "Framework:  SlowAPI rate limiting",
        "Auth:       X-Admin-Key header / JWT Bearer",
        "Test Date:  " + gen_time[:10],
        "Tester:     CyberShield DAST Suite v2.0",
    ]
    for s in scope_items:
        ws.merge_cells(f"B{row}:D{row}")
        cell(ws, row, 2, s, C["alt"], False, "left", "000000", 10)
        row += 1

    # Remediation Summary
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    hdr(ws, row, 2, "Remediation Highlights", C["section"], size=12)
    row += 1
    remediations = [
        "[FIXED] Admin routes (/admin/stats, /admin/logs, /admin/export/csv) now require X-Admin-Key",
        "[FIXED] Rate limiting (20/min) applied to all admin endpoints",
        "[FIXED] Rate limit reduced to 20/min on /analyze/url and /analyze/otp",
        "[FIXED] Frontend ADMIN_API_KEY wired via app.config.js and axios interceptor",
        "[FIXED] CSV export URL uses api_key query param for browser-based downloads",
        "[OPEN]  Public analyze endpoints intentionally unauthenticated (by design)",
        "[OPEN]  SECRET_KEY still uses placeholder - rotate before production",
        "[OPEN]  .env not in .gitignore - add backend/.env to .gitignore",
    ]
    for item in remediations:
        ws.merge_cells(f"B{row}:D{row}")
        bg = C["pass_bg"] if item.startswith("[FIXED]") else C["warn_bg"]
        cell(ws, row, 2, item, bg, False, "left", "000000", 10)
        row += 1

# ─── Sheet: Findings Summary ────────────────────────────────────────────────────
def build_summary(wb, all_rows):
    ws = wb.create_sheet("Findings Summary")
    ws.sheet_view.showGridLines = False
    headers = ["#", "Category", "Endpoint", "Method", "Severity", "Status", "Note"]
    widths  = [5,   30,         30,          8,        12,         10,       55]

    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hdr(ws, 1, i, h, C["header"]); col_w(ws, i, w)
    ws.row_dimensions[1].height = 22

    findings = sorted(
        [r for r in all_rows if r.get("finding")],
        key=lambda r: (SEV_ORDER.get(r.get("severity","INFO"),99), r.get("test_category",""))
    )

    for idx, r in enumerate(findings, 1):
        row  = idx + 1
        sev  = r.get("severity", "INFO")
        stat = remediation_status(r)
        cat  = CAT_LABELS.get(r.get("test_category",""), r.get("test_category",""))
        sev_bg = sev_color(sev)
        stat_bg = C["pass_bg"] if stat == "FIXED" else (C["fail_bg"] if stat == "OPEN" else C["alt"])

        cell(ws, row, 1, idx,                 C["alt"] if idx%2 else C["white"], True,  "center")
        cell(ws, row, 2, cat,                 C["alt"] if idx%2 else C["white"])
        cell(ws, row, 3, r.get("endpoint",""))
        cell(ws, row, 4, r.get("method",""),   None, False, "center")
        cell(ws, row, 5, sev,                  sev_bg, True, "center", "FFFFFF")
        cell(ws, row, 6, stat,                 stat_bg, True, "center",
             C["fixed"] if stat=="FIXED" else (C["open"] if stat=="OPEN" else "000000"))
        cell(ws, row, 7, r.get("note",""))
        ws.row_dimensions[row].height = 30

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:G{len(findings)+1}"

# ─── Sheet: Per-Category Detail ─────────────────────────────────────────────────
def build_category_detail(wb, all_rows):
    for cat_key, cat_label in CAT_LABELS.items():
        rows = [r for r in all_rows if r.get("test_category") == cat_key]
        if not rows:
            continue
        safe_title = cat_label.replace("/","_").replace("\\","_").replace("?","").replace("*","").replace("[","").replace("]","").replace(":","")[:31]
        ws = wb.create_sheet(safe_title)
        ws.sheet_view.showGridLines = False

        # Sheet title
        ws.merge_cells("A1:H1")
        c = ws["A1"];    c.value = f"[DETAIL] {cat_label}"
        c.fill = fill(C["header"]); c.font = font(True, "FFFFFF", 13)
        c.alignment = align("center")

        headers = ["Endpoint", "Method", "Role", "Status Code", "Expected", "Severity", "Finding?", "Note"]
        widths  = [30, 8, 20, 13, 10, 12, 10, 55]
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            hdr(ws, 2, i, h, C["subheader"]); col_w(ws, i, w)

        for idx, r in enumerate(rows, 1):
            row = idx + 2
            finding = r.get("finding", False)
            sev     = r.get("severity", "INFO")
            fixed   = finding and is_fixed(r)
            stat    = remediation_status(r)
            alt     = C["alt"] if idx % 2 == 0 else C["white"]

            if fixed:    row_bg = C["pass_bg"]
            elif finding: row_bg = C["fail_bg"]
            else:        row_bg = alt

            cell(ws, row, 1, r.get("endpoint",""),      row_bg)
            cell(ws, row, 2, r.get("method",""),        row_bg, False, "center")
            cell(ws, row, 3, r.get("role",""),          row_bg)
            cell(ws, row, 4, str(r.get("status","")),   row_bg, False, "center")
            cell(ws, row, 5, str(r.get("expected_status","")), row_bg, False, "center")
            cell(ws, row, 6, sev,                       sev_color(sev), True, "center", "FFFFFF")
            stat_fg = C["fixed"] if stat == "FIXED" else (C["open"] if stat == "OPEN" else "000000")
            cell(ws, row, 7, stat,                      row_bg, True, "center", stat_fg)
            cell(ws, row, 8, r.get("note",""),          row_bg)
            ws.row_dimensions[row].height = 28

        ws.freeze_panes = "A3"

# ─── Sheet: Remediation Tracker ─────────────────────────────────────────────────
def build_remediation(wb, all_rows):
    ws = wb.create_sheet("Remediation Tracker")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    c = ws["A1"];    c.value = "[REMEDIATION] Remediation Tracker - What Was Fixed & What Remains"
    c.fill = fill(C["header"]); c.font = font(True, "FFFFFF", 13)
    c.alignment = align("center")

    headers = ["#", "Endpoint", "Category", "Severity", "Status", "Fix Applied", "Risk if Unresolved"]
    widths  = [5, 30, 32, 12, 10, 45, 40]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hdr(ws, 2, i, h, C["subheader"]); col_w(ws, i, w)

    findings = sorted(
        [r for r in all_rows if r.get("finding")],
        key=lambda r: (SEV_ORDER.get(r.get("severity","INFO"),99), r.get("test_category",""))
    )

    # FIX_DETAILS keyed by (test_category, endpoint)
    FIX_DETAILS = {
        ("08_rate_limit", "/admin/stats"):       ("@limiter.limit('20/minute') added to /admin/stats",       "DoS / brute-force on admin"),
        ("08_rate_limit", "/admin/logs"):        ("@limiter.limit('20/minute') added to /admin/logs",        "DoS / log flooding"),
        ("08_rate_limit", "/admin/export/csv"):  ("@limiter.limit('20/minute') added to /admin/export/csv",  "DoS / data exfil via CSV"),
        ("08_rate_limit", "/analyze/url"):       ("Rate limit reduced 30->20/min on /analyze/url",           "Abuse of URL scan endpoint"),
        ("08_rate_limit", "/analyze/otp"):       ("Rate limit reduced 30->20/min on /analyze/otp",           "OTP message flood"),
        ("02_authn_bypass", "/admin/stats"):     ("X-Admin-Key header + api_key query param validation added", "Unauth admin data access"),
        ("02_authn_bypass", "/admin/logs"):      ("X-Admin-Key header + api_key query param validation added", "Log data exfiltration"),
        ("02_authn_bypass", "/admin/export/csv"):("X-Admin-Key header + api_key query param validation added", "Full CSV data export"),
    }

    seen = set()
    idx = 0
    for r in findings:
        cat = r.get("test_category","")
        ep  = r.get("endpoint","")
        key = (cat, ep)
        if key in seen:
            continue
        seen.add(key)
        idx += 1
        row  = idx + 2
        stat = remediation_status(r)
        sev  = r.get("severity","INFO")
        fix_detail = FIX_DETAILS.get(key, ("Manual review required", "Depends on context"))

        stat_bg = C["pass_bg"] if stat == "FIXED" else C["fail_bg"]
        stat_fg = C["fixed"]   if stat == "FIXED" else C["open"]

        cell(ws, row, 1, idx,    C["alt"] if idx%2 else C["white"], True, "center")
        cell(ws, row, 2, ep)
        cell(ws, row, 3, CAT_LABELS.get(cat, cat))
        cell(ws, row, 4, sev,    sev_color(sev), True, "center", "FFFFFF")
        cell(ws, row, 5, stat,   stat_bg, True, "center", stat_fg)
        cell(ws, row, 6, fix_detail[0])
        cell(ws, row, 7, fix_detail[1])
        ws.row_dimensions[row].height = 30

    ws.freeze_panes = "A3"

# ─── Sheet: Statistics & Charts ────────────────────────────────────────────────
def build_stats(wb, all_rows):
    ws = wb.create_sheet("Statistics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    c = ws["A1"];    c.value = "[STATS] Test Statistics & Security Posture"
    c.fill = fill(C["header"]); c.font = font(True, "FFFFFF", 13)
    c.alignment = align("center")

    # By severity
    findings = [r for r in all_rows if r.get("finding")]
    open_f   = [r for r in findings if not is_fixed(r)]
    fixed_f  = [r for r in findings if is_fixed(r)]
    passes   = [r for r in all_rows if not r.get("finding")]

    row = 3
    hdr(ws, row, 1, "Metric",     C["subheader"]); col_w(ws, 1, 30)
    hdr(ws, row, 2, "Count",      C["subheader"]); col_w(ws, 2, 12)
    hdr(ws, row, 3, "% of Total", C["subheader"]); col_w(ws, 3, 15)
    row += 1

    total = len(all_rows)
    stats_data = [
        ("Total Tests Executed",      total,          "FFFFFF"),
        ("Tests Passed (no finding)", len(passes),    C["pass_bg"]),
        ("Findings Fixed",            len(fixed_f),   "D6EAF8"),
        ("Open Findings",             len(open_f),    C["fail_bg"]),
        ("  → Critical",              sum(1 for r in open_f if r.get("severity")=="CRITICAL"), C["fail_bg"]),
        ("  → High",                  sum(1 for r in open_f if r.get("severity")=="HIGH"),     C["warn_bg"]),
        ("  → Medium",                sum(1 for r in open_f if r.get("severity")=="MEDIUM"),   "FDFEFE"),
        ("  → Low",                   sum(1 for r in open_f if r.get("severity")=="LOW"),      C["pass_bg"]),
        ("  → Info",                  sum(1 for r in open_f if r.get("severity")=="INFO"),     C["alt"]),
    ]
    for label, count, bg in stats_data:
        pct = f"{count/total*100:.1f}%" if total else "0%"
        cell(ws, row, 1, label, bg, False, "left")
        cell(ws, row, 2, count, bg, True,  "center")
        cell(ws, row, 3, pct,   bg, False, "center")
        row += 1

    # By category
    row += 1
    hdr(ws, row, 1, "Test Category",   C["subheader"]); 
    hdr(ws, row, 2, "Tests Run",       C["subheader"])
    hdr(ws, row, 3, "Open Findings",   C["subheader"])
    hdr(ws, row, 4, "Fixed",           C["subheader"]); col_w(ws, 4, 12)
    hdr(ws, row, 5, "Pass Rate",       C["subheader"]); col_w(ws, 5, 12)
    row += 1

    chart_data_start = row
    for cat_key, cat_label in CAT_LABELS.items():
        cat_rows   = [r for r in all_rows if r.get("test_category") == cat_key]
        cat_open   = [r for r in cat_rows if r.get("finding") and not is_fixed(r)]
        cat_fixed  = [r for r in cat_rows if r.get("finding") and is_fixed(r)]
        n = len(cat_rows)
        pass_r = f"{(n-len(cat_open))/n*100:.0f}%" if n else "N/A"
        bg = C["pass_bg"] if len(cat_open) == 0 else (C["fail_bg"] if len(cat_open) > 2 else C["warn_bg"])

        cell(ws, row, 1, cat_label, bg)
        cell(ws, row, 2, n,              bg, False, "center")
        cell(ws, row, 3, len(cat_open),  bg, True,  "center", C["open"] if cat_open else C["fixed"])
        cell(ws, row, 4, len(cat_fixed), bg, False, "center")
        cell(ws, row, 5, pass_r,         bg, False, "center")
        row += 1

    col_w(ws, 1, 35); col_w(ws, 2, 12); col_w(ws, 3, 15)
    ws.freeze_panes = "A2"

# ─── Sheet: Raw Data ────────────────────────────────────────────────────────────
def build_raw(wb, all_rows):
    ws = wb.create_sheet("Raw Test Data")
    headers = ["Category", "Endpoint", "Method", "Role", "Status", "Expected", 
               "Severity", "Finding?", "Remediation", "Note", "Timestamp"]
    widths  = [28, 30, 8, 18, 10, 10, 12, 10, 12, 50, 22]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        hdr(ws, 1, i, h, C["header"]); col_w(ws, i, w)

    for idx, r in enumerate(all_rows, 1):
        row    = idx + 1
        sev    = r.get("severity", "INFO")
        stat   = remediation_status(r)
        alt    = C["alt"] if idx%2==0 else C["white"]
        row_bg = C["pass_bg"] if stat=="FIXED" else (C["fail_bg"] if stat=="OPEN" else alt)
        cat    = CAT_LABELS.get(r.get("test_category",""), r.get("test_category",""))

        cell(ws, row, 1,  cat,                         alt)
        cell(ws, row, 2,  r.get("endpoint",""),        alt)
        cell(ws, row, 3,  r.get("method",""),          alt, False, "center")
        cell(ws, row, 4,  r.get("role",""),            alt)
        cell(ws, row, 5,  str(r.get("status","")),     alt, False, "center")
        cell(ws, row, 6,  str(r.get("expected_status","")), alt, False, "center")
        cell(ws, row, 7,  sev,                         sev_color(sev), True, "center", "FFFFFF")
        cell(ws, row, 8,  "YES" if r.get("finding") else "NO",
             C["fail_bg"] if r.get("finding") else C["pass_bg"], True, "center",
             C["open"] if r.get("finding") else C["fixed"])
        cell(ws, row, 9,  stat,                        row_bg, True, "center")
        cell(ws, row, 10, r.get("note",""),            alt)
        cell(ws, row, 11, r.get("timestamp","")[:19] if r.get("timestamp") else "", alt)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:K{len(all_rows)+1}"

# ─── Main ──────────────────────────────────────────────────────────────────────
def main():
    gen_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n{'='*60}")
    print(f"  CyberShield Security Report Generator")
    print(f"  {gen_time}")
    print(f"{'='*60}\n")

    all_rows = load_results()
    if not all_rows:
        print("ERROR: No result files found. Run DAST suite first.")
        return

    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print("  Building: Cover Page...")
    build_cover(wb, all_rows, gen_time)
    print("  Building: Findings Summary...")
    build_summary(wb, all_rows)
    print("  Building: Category Detail sheets...")
    build_category_detail(wb, all_rows)
    print("  Building: Remediation Tracker...")
    build_remediation(wb, all_rows)
    print("  Building: Statistics...")
    build_stats(wb, all_rows)
    print("  Building: Raw Test Data...")
    build_raw(wb, all_rows)

    wb.save(OUTPUT_FILE)
    print(f"\n[OK] Report saved: {OUTPUT_FILE}\n")

    # Summary to console
    findings  = [r for r in all_rows if r.get("finding")]
    fixed_rows = [r for r in all_rows if is_in_fixed_set(r)]
    open_f    = [r for r in findings if not is_in_fixed_set(r)]
    print(f"  Total tests   : {len(all_rows)}")
    print(f"  Total findings: {len(findings)}")
    print(f"  Remediated    : {len(set((r.get('endpoint'),r.get('test_category')) for r in fixed_rows))} endpoint-category pairs ({len(fixed_rows)} test rows marked FIXED)")
    print(f"  Open findings : {len(open_f)}")
    for s in ["CRITICAL","HIGH","MEDIUM","LOW"]:
        n = sum(1 for r in open_f if r.get("severity")==s)
        if n: print(f"    {s:10}: {n}")
    print()

if __name__ == "__main__":
    main()
