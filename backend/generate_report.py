"""
CyberShield Security Assessment — Excel Report Generator
Run: python generate_report.py
Output: automated_test/CyberShield_Security_Report.xlsx
Requires: pip install openpyxl
"""

import json
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    print("Installing openpyxl...")
    os.system("pip install openpyxl")
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

# ─── Color Palette ────────────────────────────────────────────────────────────
COLORS = {
    "CRITICAL": "C0392B",
    "HIGH":     "E67E22",
    "MEDIUM":   "F1C40F",
    "LOW":      "27AE60",
    "INFO":     "2980B9",
    "header_bg":"1A1A2E",
    "header_fg":"FFFFFF",
    "section":  "16213E",
    "alt_row":  "F2F2F2",
    "white":    "FFFFFF",
    "green":    "27AE60",
    "red":      "C0392B",
    "border":   "CCCCCC",
}

SEVERITY_ORDER = {"CRITICAL": 0, "HIGH": 1, "MEDIUM": 2, "LOW": 3, "INFO": 4}

def cell_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font(bold=True, color="FFFFFF", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def normal_font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def center_align(wrap=True):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left_align(wrap=True):
    return Alignment(horizontal="left", vertical="top", wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=COLORS["border"])
    return Border(left=s, right=s, top=s, bottom=s)

def set_header(ws, row, col, text, bg=COLORS["header_bg"], fg="FFFFFF", size=11, bold=True):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = cell_fill(bg)
    c.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    c.alignment = center_align()
    c.border = thin_border()
    return c

def set_cell(ws, row, col, value, bg=None, bold=False, align="left", color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    if bg:
        c.fill = cell_fill(bg)
    c.font = Font(bold=bold, color=color, size=10, name="Calibri")
    c.alignment = left_align() if align == "left" else center_align()
    c.border = thin_border()
    return c

# ─── Load data ─────────────────────────────────────────────────────────────────
base = os.path.join(os.path.dirname(__file__), "automated_test")

with open(os.path.join(base, "report.json")) as f:
    report = json.load(f)
with open(os.path.join(base, "endpoint_inventory.json")) as f:
    inventory = json.load(f)
with open(os.path.join(base, "auth_tests.json")) as f:
    auth = json.load(f)
with open(os.path.join(base, "authz_tests.json")) as f:
    authz = json.load(f)
with open(os.path.join(base, "validation_tests.json")) as f:
    validation = json.load(f)
with open(os.path.join(base, "rate_limit_tests.json")) as f:
    rate = json.load(f)
with open(os.path.join(base, "secrets_scan.json")) as f:
    secrets = json.load(f)

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.create_sheet("Executive Summary")
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 30
ws1.column_dimensions["B"].width = 50
ws1.column_dimensions["C"].width = 18

# Title
ws1.merge_cells("A1:C1")
t = ws1["A1"]
t.value = "🛡️  CyberShield API — Security Assessment Report"
t.fill = cell_fill(COLORS["header_bg"])
t.font = Font(bold=True, color="FFFFFF", size=16, name="Calibri")
t.alignment = center_align()
ws1.row_dimensions[1].height = 40

# Subtitle
ws1.merge_cells("A2:C2")
s = ws1["A2"]
s.value = f"Assessment Date: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Method: Static Code Analysis (SAST) | Scope: Full Backend"
s.fill = cell_fill(COLORS["section"])
s.font = Font(color="AAAAAA", size=10, name="Calibri", italic=True)
s.alignment = center_align()

ws1.row_dimensions[3].height = 10

# Key Metrics
es = report["executive_summary"]
metrics = [
    ("Endpoints Discovered",  str(es["total_endpoints_discovered"]),  COLORS["INFO"]),
    ("Total Tests Run",       str(es["total_tests_run"]),             COLORS["INFO"]),
    ("Total Findings",        str(es["total_findings"]),              COLORS["HIGH"]),
    ("Overall Risk Rating",   es["overall_risk_rating"],              COLORS["CRITICAL"]),
]
set_header(ws1, 4, 1, "Metric", COLORS["section"])
set_header(ws1, 4, 2, "Value", COLORS["section"])
set_header(ws1, 4, 3, "Indicator", COLORS["section"])
ws1.row_dimensions[4].height = 22

for i, (label, val, color) in enumerate(metrics, 5):
    bg = COLORS["alt_row"] if i % 2 == 0 else COLORS["white"]
    set_cell(ws1, i, 1, label, bg=bg, bold=True)
    set_cell(ws1, i, 2, val, bg=bg)
    c = ws1.cell(row=i, column=3, value="●")
    c.fill = cell_fill(bg)
    c.font = Font(color=color, size=14, bold=True)
    c.alignment = center_align()
    c.border = thin_border()
    ws1.row_dimensions[i].height = 20

# Severity breakdown
ws1.row_dimensions[10].height = 10
set_header(ws1, 11, 1, "Severity", COLORS["section"])
set_header(ws1, 11, 2, "Count", COLORS["section"])
set_header(ws1, 11, 3, "Risk Level", COLORS["section"])
ws1.row_dimensions[11].height = 22

sev_map = es["findings_by_severity"]
for i, sev in enumerate(["CRITICAL","HIGH","MEDIUM","LOW","INFO"], 12):
    count = sev_map.get(sev, 0)
    bg = COLORS["alt_row"] if i % 2 == 0 else COLORS["white"]
    c1 = ws1.cell(row=i, column=1, value=sev)
    c1.fill = cell_fill(COLORS[sev] if sev in COLORS else COLORS["INFO"])
    c1.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    c1.alignment = center_align()
    c1.border = thin_border()
    set_cell(ws1, i, 2, count, bg=bg, align="center")
    bar = "█" * count
    c3 = ws1.cell(row=i, column=3, value=bar)
    c3.fill = cell_fill(bg)
    c3.font = Font(color=COLORS[sev] if sev in COLORS else COLORS["INFO"], size=10)
    c3.alignment = left_align(wrap=False)
    c3.border = thin_border()
    ws1.row_dimensions[i].height = 20

# Summary text
ws1.row_dimensions[18].height = 10
ws1.merge_cells("A19:C25")
summ = ws1["A19"]
summ.value = "EXECUTIVE SUMMARY\n\n" + es["summary"]
summ.fill = cell_fill("FFF9E6")
summ.font = Font(size=10, name="Calibri", color="1A1A2E")
summ.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
summ.border = thin_border()
ws1.row_dimensions[19].height = 100

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ENDPOINT INVENTORY
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Endpoint Inventory")
ws2.sheet_view.showGridLines = False
cols2 = ["ID","Method","Path","Auth Required","Auth Mechanism","Expected Roles","Description","Router File"]
widths2 = [6, 10, 30, 16, 20, 18, 50, 25]
for i, (h, w) in enumerate(zip(cols2, widths2), 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
    set_header(ws2, 1, i, h)
ws2.row_dimensions[1].height = 24

METHOD_COLORS = {"GET":"27AE60","POST":"2980B9","DELETE":"C0392B","PUT":"E67E22","PATCH":"8E44AD"}

for r, ep in enumerate(inventory["endpoints"], 2):
    bg = COLORS["alt_row"] if r % 2 == 0 else COLORS["white"]
    vals = [
        ep["id"], ep["method"], ep["path"],
        "Yes" if ep["auth_required"] else "No ⚠",
        ep["auth_mechanism"],
        ", ".join(ep["expected_roles"]),
        ep["description"],
        ep["router_file"],
    ]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(row=r, column=c, value=v)
        cell.fill = cell_fill(bg)
        cell.font = normal_font()
        cell.alignment = left_align()
        cell.border = thin_border()
        if c == 2:  # method column
            mc = METHOD_COLORS.get(v, "666666")
            cell.fill = cell_fill(mc)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = center_align()
        if c == 4 and v.startswith("No"):
            cell.fill = cell_fill("FDECEA")
            cell.font = Font(bold=True, color=COLORS["CRITICAL"], size=10, name="Calibri")
    ws2.row_dimensions[r].height = 40

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — ALL FINDINGS (Master Report)
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("All Findings")
ws3.sheet_view.showGridLines = False
ws3.freeze_panes = "A3"

cols3 = ["#","Finding ID","Title","Endpoint","Method","Severity","CVSS","Category","Expected","Finding","Remediation","Priority"]
widths3 = [5, 12, 40, 30, 10, 12, 8, 25, 20, 10, 60, 10]
for i, (h, w) in enumerate(zip(cols3, widths3), 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

ws3.merge_cells("A1:L1")
title3 = ws3["A1"]
title3.value = "CyberShield — Complete Findings Register"
title3.fill = cell_fill(COLORS["header_bg"])
title3.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
title3.alignment = center_align()
ws3.row_dimensions[1].height = 30

for i, h in enumerate(cols3, 1):
    set_header(ws3, 2, i, h)
ws3.row_dimensions[2].height = 22

findings = sorted(report["findings"], key=lambda x: SEVERITY_ORDER.get(x["severity"], 9))

for r, f in enumerate(findings, 3):
    sev = f["severity"]
    bg_sev = COLORS.get(sev, "AAAAAA")
    row_bg = COLORS["alt_row"] if r % 2 == 0 else COLORS["white"]

    vals = [
        r - 2,
        f["finding_id"],
        f["title"],
        f["endpoint"],
        f["method"],
        f["severity"],
        f.get("cvss_estimate", "N/A"),
        f["test_category"],
        f.get("expected_status", ""),
        "✗ FINDING" if f["finding"] else "✓ PASS",
        f["remediation"],
        "",  # priority filled below
    ]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=r, column=c, value=v)
        cell.fill = cell_fill(row_bg)
        cell.font = normal_font()
        cell.alignment = left_align()
        cell.border = thin_border()

        if c == 6:  # severity
            cell.fill = cell_fill(bg_sev)
            cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
            cell.alignment = center_align()
        if c == 10:  # finding
            if f["finding"]:
                cell.fill = cell_fill("FDECEA")
                cell.font = Font(bold=True, color=COLORS["CRITICAL"], size=10, name="Calibri")
            else:
                cell.fill = cell_fill("E9F7EF")
                cell.font = Font(bold=True, color=COLORS["green"], size=10, name="Calibri")
            cell.alignment = center_align()
        if c == 7:  # cvss
            cell.alignment = center_align()
    ws3.row_dimensions[r].height = 60

# Fill priority from remediation_priority
priority_map = {}
for p in report["remediation_priority"]:
    for fid in p["finding_ids"]:
        priority_map[fid] = p["priority"]

for r, f in enumerate(findings, 3):
    pval = priority_map.get(f["finding_id"], "-")
    cell = ws3.cell(row=r, column=12, value=f"P{pval}" if pval != "-" else "-")
    cell.alignment = center_align()
    cell.border = thin_border()
    if isinstance(pval, int) and pval <= 2:
        cell.fill = cell_fill("FDECEA")
        cell.font = Font(bold=True, color=COLORS["CRITICAL"], size=10, name="Calibri")
    elif isinstance(pval, int) and pval <= 4:
        cell.fill = cell_fill("FEF9E7")
        cell.font = Font(bold=True, color=COLORS["HIGH"], size=10, name="Calibri")
    else:
        cell.fill = cell_fill(COLORS["alt_row"])
        cell.font = normal_font()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — AUTH TESTS
# ══════════════════════════════════════════════════════════════════════════════
def build_test_sheet(wb, sheet_name, test_data):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    ws.merge_cells("A1:J1")
    title = ws["A1"]
    title.value = f"{sheet_name} — {test_data.get('summary', '')}"
    title.fill = cell_fill(COLORS["header_bg"])
    title.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    title.alignment = center_align()
    ws.row_dimensions[1].height = 28

    headers = ["Test ID","Endpoint","Method","Role","Category","Severity","Finding","Expected","Evidence / Actual Behavior","Notes"]
    widths  = [12, 30, 10, 20, 25, 12, 12, 20, 55, 50]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        set_header(ws, 2, i, h)
    ws.row_dimensions[2].height = 22

    tests = test_data.get("tests", [])
    for r, t in enumerate(tests, 3):
        row_bg = COLORS["alt_row"] if r % 2 == 0 else COLORS["white"]
        sev = t.get("severity", "INFO")
        vals = [
            t.get("test_id",""),
            t.get("endpoint",""),
            t.get("method",""),
            t.get("role",""),
            t.get("category",""),
            sev,
            "✗ FINDING" if t.get("finding") else "✓ PASS",
            str(t.get("expected_status", t.get("expected", ""))),
            t.get("evidence", t.get("actual_behavior", "")),
            t.get("notes",""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.fill = cell_fill(row_bg)
            cell.font = normal_font()
            cell.alignment = left_align()
            cell.border = thin_border()
            if c == 6:
                cell.fill = cell_fill(COLORS.get(sev, "AAAAAA"))
                cell.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                cell.alignment = center_align()
            if c == 7:
                if t.get("finding"):
                    cell.fill = cell_fill("FDECEA")
                    cell.font = Font(bold=True, color=COLORS["CRITICAL"], size=10)
                else:
                    cell.fill = cell_fill("E9F7EF")
                    cell.font = Font(bold=True, color=COLORS["green"], size=10)
                cell.alignment = center_align()
        ws.row_dimensions[r].height = 65
    return ws

build_test_sheet(wb, "Auth Tests", auth)
build_test_sheet(wb, "AuthZ Tests", authz)
build_test_sheet(wb, "Validation Tests", validation)
build_test_sheet(wb, "Rate Limit Tests", rate)
build_test_sheet(wb, "Secrets Scan", secrets)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8 — REMEDIATION ROADMAP
# ══════════════════════════════════════════════════════════════════════════════
ws8 = wb.create_sheet("Remediation Roadmap")
ws8.sheet_view.showGridLines = False
ws8.column_dimensions["A"].width = 10
ws8.column_dimensions["B"].width = 20
ws8.column_dimensions["C"].width = 50
ws8.column_dimensions["D"].width = 15
ws8.column_dimensions["E"].width = 25

ws8.merge_cells("A1:E1")
t8 = ws8["A1"]
t8.value = "CyberShield — Remediation Roadmap (Prioritized)"
t8.fill = cell_fill(COLORS["header_bg"])
t8.font = Font(bold=True, color="FFFFFF", size=14, name="Calibri")
t8.alignment = center_align()
ws8.row_dimensions[1].height = 30

for i, h in enumerate(["Priority","Finding IDs","Action","Effort","Status"], 1):
    set_header(ws8, 2, i, h)
ws8.row_dimensions[2].height = 22

priority_colors = {1: COLORS["CRITICAL"], 2: COLORS["CRITICAL"], 3: COLORS["HIGH"],
                   4: COLORS["HIGH"], 5: COLORS["MEDIUM"], 6: COLORS["LOW"], 7: COLORS["LOW"]}

for r, p in enumerate(report["remediation_priority"], 3):
    row_bg = COLORS["alt_row"] if r % 2 == 0 else COLORS["white"]
    pnum = p["priority"]
    pc = priority_colors.get(pnum, COLORS["INFO"])

    c1 = ws8.cell(row=r, column=1, value=f"P{pnum}")
    c1.fill = cell_fill(pc)
    c1.font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
    c1.alignment = center_align()
    c1.border = thin_border()

    set_cell(ws8, r, 2, ", ".join(p["finding_ids"]), bg=row_bg, bold=True)
    set_cell(ws8, r, 3, p["action"], bg=row_bg)
    set_cell(ws8, r, 4, p["effort"], bg=row_bg, align="center")
    cell_status = ws8.cell(row=r, column=5, value="⬜ Open")
    cell_status.fill = cell_fill(row_bg)
    cell_status.font = Font(color="666666", size=10, name="Calibri")
    cell_status.alignment = center_align()
    cell_status.border = thin_border()
    ws8.row_dimensions[r].height = 45

# ─── Save ──────────────────────────────────────────────────────────────────────
out_path = os.path.join(base, "CyberShield_Security_Report.xlsx")
wb.save(out_path)
print(f"\n[OK] Report saved to: {out_path}")
print(f"    Sheets: Executive Summary | Endpoint Inventory | All Findings")
print(f"            Auth Tests | AuthZ Tests | Validation Tests | Rate Limit Tests | Secrets Scan | Remediation Roadmap")
print(f"\n[!] CRITICAL: {es['findings_by_severity']['CRITICAL']} critical findings -- see Remediation Roadmap sheet for priorities.")
